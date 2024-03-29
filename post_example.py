# PYTHONDA QIZIQARLI POSTLAR

# def key_values(**kwargs):
#     for key, value in kwargs.items():
#         print(f"Key: {key}\nValue: {value}")

# key_values(first_name="Erali", last_name="Abdinazarov")

# # *args
# def func(arg1, arg2, *args):

#     print("first arg: {}".format(arg1))
#     print("second arg: {}".format(arg2))

#     for arg in args:
#         print(arg)

# func(1,2,3,4,5)

# def args_kwargs(*args, **kwargs):
#     print("args: ", args)
#     print("kwargs: ", kwargs)

# args_kwargs(1, 2, 3, first="birinchi", middle="o'rta", last="oxirgi")


# # Object *args
# class car:
#     def __init__(self, *args):
#         self.speed = args[0]
#         self.color = args[1]

#     def get_car_info(self):
#         return f"Car color: {self.color}\nCar speed: {self.speed}"

# audi = car(200, 'red')
# bmw = car(250, 'black')

# print("Audi info:")
# print(audi.get_car_info(), "\n")
# print("BMW info:")
# print(bmw.get_car_info())

# # Object **kwargs
# class car:
#     def __init__(self, **kwargs):
#         self.speed = kwargs['sp']
#         self.color = kwargs['col']

#     def get_car_info(self):
#         return f"Car color: {self.color}\nCar speed: {self.speed}"

# audi = car(sp = 200, col = 'red')
# bmw = car(sp = 250, col = 'black')

# print("Audi info:")
# print(audi.get_car_info(), "\n")
# print("BMW info:")
# print(bmw.get_car_info())

# import phonenumbers
# from phonenumbers import geocoder, carrier

# phone_number = input("Enter your Phone Number: ")
# phoneNumber = phonenumbers.parse(phone_number)
# region = geocoder.description_for_number(phoneNumber, "uz")

# print(region)

# def filter_list(list, number):
#     if number != 0:
#         for i in list:
#             if i % 2 == 1:
#                 list.remove(i)
#     else:
#         for i in list:
#             if i % 2 == 0:
#                 list.remove(i)
#     print(list)
# filter_list([1,2,3,4,5,6,7,8,9], 1)
# filter_list([1,2,3,4,5,6,7,8,9], 0)

# lambda, filter
# a = [100, 2, 8, 60, 5, 4, 3, 31, 10, 11]
# filtered = list(filter(lambda x: x % 2 == 0, a))
# print(filtered)

# def power(n):

#     return lambda a: a ** n

# lmdb = power(2)
# print(lmdb(3))

# # nonlocal
# def external_func():
#   name = "Alijon"

#   def internal_func():
#     nonlocal name
#     name = "Valijon"

#   internal_func() 

#   return name

# print("Name:",external_func())

# # global
# def global_func():
#     global name
#     name = "John Doe"
#     return name

# print(global_func())
# print(name)


# class Person:
#     def __init__(self, first_name, last_name, age):
#         self.first_name = first_name
#         self.last_name = last_name
#         self.age = age

#     def full_name(self):
#         return f"{self.first_name} {self.last_name}"

# person = Person("Erali", "Abdinazarov", 22)
# full = person.full_name()
# print(full)

# class Student(Person):
#     def __init__(self, first_name, last_name, age, job):
#         super().__init__(first_name, last_name, age)
#         self.job = job

#     def full_name(self):
#         return super().full_name()

#     def info(self):
#         return f"Ismim {self.first_name} {self.last_name}, yoshim {self.age} da va kasbim {self.job} 👍"

# student = Student("Erali", "Abdinazarov", 22, "Dasturchi")
# print(student.full_name())
# print(student.info())


# n ta elementga ega massiv va k butun son berilgan. Massivning har bir elementini a[k] ga
# orttiruvchi programma tuzilsin. (1 <= k <= n)

# a = list(map(int, input("n ta son kiriting: ").split()))
# k = int(input("k ni kiriting: "))

# a_k = []

# for i in a:
#     l = i + a[k]
#     a_k.append(l)
# print(a_k)

# Array 83
# a = list(map(int, input("n ta son kiriting: ").split()))
# # k = int(input("k ni kiriting: "))
# ab = []

# for i in range(len(a)):
#     ab.append(a[i])
# print(ab)

# n = int(input("Son kiriting: "))
# a = []
# for i in range(n+1):
#     a.append(int(input("a["+str(i)+"]=")))
# # print(a)
# for j in range(len(a)):
#     a.pop(j)
# print(a)

# # Bubble Sort
# import random
# import time

# nums = list(range(10))
# random.shuffle(nums)
# print(f"Tartibsiz ro'yxat: {nums}")

# start_time = time.time()
# def bubble(nums):
#     n = len(nums)
#     swaps_count = 0

#     for i in range(n - 1):
#         if nums[i] > nums[i+1]:
#             nums[i], nums[i+1] = nums[i+1], nums[i]
#             swaps_count += 1

#     return swaps_count
# bubble(nums)

# def bubble_sort(nums):
#     n = len(nums)

#     for _ in range(n-1):
#         if bubble(nums) == 0:
#             break

# bubble_sort(nums)

# end_time = time.time()
# timer = end_time - start_time
# print(f"Bubble sort algoritmiga ketgan vaqt: {timer}")

# print(f"Tartiblangan ro'yxat: {nums}")


# Selection Sort
nums = [8, 5, 6, 7, 4, 2, 3]

def select(nums, start):
    n = len(nums)
    pos = start
    for i in range(start, n):
        if nums[i] < nums[pos]:
            pos = i
    return pos

select(nums, 0)

def selection_sort(nums):
    n = len(nums)
    for i in range(n):
        pos = select(nums, i)
        nums[pos], nums[i] = nums[i], nums[pos]
        print(nums)

selection_sort(nums)

def split_and_join(line):
    line_split = line.split(' ')
    t = "-".join(line_split)
    
    print(t)

split_and_join("this is a string")

# biron harf o'rniga boshqa bir harf qo'yish
def mutate_string(string, position, character):
    string_list = list(string)
    for i in range(len(string_list)):
        if int(position) == i:
            del string_list[i]
            string_list.insert(int(position), character)
    string = ''.join(string_list)
    return string

if __name__ == '__main__':
    s = input("Matn kiriting: ")
    i, c = input("index raqam probel tashlab belgi kiriting: ").split()
    s_new = mutate_string(s, int(i), c)
    print(s_new)


def print_formatted(number):
    if number >= 1 and number <= 99:
        for i in range(1, number+1):
            decimal = i
            octal = oct(i)
            hexadecimal = hex(i)
            binary = bin(i)

            # a = len(str(decimal))
            # b = len(octal[2:])
            # c = len(hexadecimal[2:])
            # d = len(binary[2:])

            result = "{0:>2} {1:>3} {2:>2} {3:>7}".format(decimal, octal[2:], hexadecimal[2:].title(), binary[2:])
            print(result)

if __name__ == '__main__':
    n = int(input())
    print_formatted(n)

a, b = map(int, input().split())
middle_arifmetik = (a+b)/2
middle_geometrik = (a*b)**(1/2)
if middle_arifmetik > middle_geometrik:
    print(">")
elif middle_geometrik > middle_arifmetik:
    print("<")
else:
    print("=")

n = int(input())
s = 0
if n > 0:
    for i in range(1, n+1):
        s += i
if n < 0:
    for i in range(n, 2):
        s += i
print(s)

n = int(input())
list_sort = []
for i in range(n):
    m = int(input())
    list_sort.append(m)

def sort(list_sort):
    for i in range(len(list_sort)-1):
        if list_sort[i] > list_sort[i+1]:
            list_sort[i], list_sort[i+1] = list_sort[i+1], list_sort[i]

sort(list_sort)

def sort1(list_sort):
    for j in range(len(list_sort)-1):
        if sort(list_sort) == 0:
            break

sort1(list_sort)
for i in list_sort:
        print(i)


if __name__ == '__main__':
    s = input()
    if s.isalnum():
        print("True")
    elif s.isalpha():
        print("True")
    elif s.isdigit():
        print("True")
    elif s.islower():
        print("True")
    elif s.isupper():
        print("True")
    else:
        print("False")

    lst = list(s)
    print(lst.isalnum())
    print(lst.isalpha())
    print(lst.isdigit())
    print(lst.islower())
    print(lst.isupper())

    print(bool(re.findall("[a-zA-Z0-9+]", s)))
    print(bool(re.findall("[a-zA-Z]", s)))
    print(bool(re.findall("[a-z0-9]", s)))
    print(bool(re.findall("[a-z+]", s)))
    print(bool(re.findall("A-Z+]", s)))

import textwrap

def wrap(string, max_width):
    return textwrap.fill(string, width=max_width)

if __name__ == '__main__':
    string, max_width = input(), int(input())
    result = wrap(string, max_width)
    print(result)

# Berilgan sonni teskari tartibda chiqarish
def print_digit(num):
    """
        Kod bu yerga yozing!
    """
    if num == 0:
        print(0)
        return
    while num != 0:
        digit = num % 10
        print(digit)
        num //=10
number = int(input("Sonni kiriting: "))
print_digit(num=number)

# Berilgan sonni teskari bo'lgan sonni qaytaring
def reverse_digit(num):
    """
        kodni shu yerga yozing!
    """
    result = 0
    while num != 0:
        result *= 10
        digit = num % 10
        result += digit
        num //= 10
    print(result)
number = int(input("Son kiriting: "))
reverse_digit(num=number)

def reverse_digit(num):
    """
        kodni shu yerga yozing!
    """
    result = ''
    while num != 0:
        digit = num % 10
        result += str(digit)
        num //= 10
    print(result)
number = int(input("Son kiriting: "))
reverse_digit(num=number)

# Berilgan sonni ikkilik sanoq tizimiga o'tkazish
def reverse_digit(num: int) -> str:
    """
        kodni shu yerga yozing!
    """
    result = ''
    while num != 0:
        digit = num % 2
        result += str(digit)
        num //= 2
    print(result)

    
number = int(input("Son kiriting: "))
reverse_digit(num=number)

# Berilgan son ikkining darjasi ekanini tekshiring
def reverse_digit(num: int) -> bool:
    """
    kodni shu yerga yozing!
    """
    bit_count = 0
    while num != 0:
        bit_count += num & 1
        print("bir",bit_count)
        num = num >> 1
        print("ikki", num)
    print(bit_count == 1)

    
number = int(input("Son kiriting: "))
reverse_digit(num=number)


from math import factorial
class Solution:
    def solve(self, n):
        return factorial(n)

number = int(input("Sonni kiriting: "))
solution = Solution()
result = solution.solve(number)
print(result)


# 1 dan N gacha bo'lgan toq sonlar yig'indisini toping O(1) vaqtda
def sum_of_add_number(n: int) -> int:
    """
    Kodni shu yerga yozing!
    """
    s = 0
    for i in range(1, n, 2):
        s += i
    print(s)
sum_of_add_number(10)
end_time = time.time()

# Palindrom so'zlar
def is_palindrome(word: str) -> bool:
    low, high = 0, len(word) - 1

    while low < high:
        # print(word[low], word[high])
        if word[low] != word[high]:
            return False
        low += 1
        high -= 1
    return True
print(is_palindrome("kiyik"))


if __name__ == '__main__':
    n = int(input())
    student_marks = {}
    for _ in range(n):
        name, *line = input().split()
        scores = list(map(float, line))
        student_marks[name] = scores
    query_name = input()

    s=0
    for i in student_marks[query_name]:
        s+=i
    midd = s/len(student_marks[query_name])
    print(f"{midd:.2f}")


from fractions import Fraction
from functools import reduce

def product(fracs):
    t = reduce(lambda a, b: a*b, fracs) # complete this line with a reduce statement
    return t.numerator, t.denominator

if __name__ == '__main__':
    fracs = []
    for _ in range(int(input())):
        fracs.append(Fraction(*map(int, input().split())))
    result = product(fracs)
    print(*result)

from functools import reduce
 
lst = [1, 3, 5, 6, 2]

sum_elements = reduce(lambda a, b: a+b, lst)
print("Ro'yxat elementlarining yig'indisi:",sum_elements)

max_element = reduce(lambda a, b: a if a > b else b, lst)
print("Ro'yxatdagi eng katta element:",max_element)


from fractions import Fraction
  
print (Fraction(11, 35))
  
print (Fraction(10, 18))

print (Fraction(1, 0))

print (Fraction(1.13))


def count_substring(string, sub_string):
    k = 0
    j = len(sub_string)
    for i in range(0, len(string)):
        # print(string[i:j])
        if string[i:j] == sub_string:
            k += 1
        j += 1
    return k

if __name__ == '__main__':
    string = input().strip()
    sub_string = input().strip()
    
    count = count_substring(string, sub_string)
    print(count)



def swap_case(s):
    return s.swapcase()

if __name__ == '__main__':
    s = input()
    result = swap_case(s)
    print(result)



a = int(input())
b = int(input())
print(a//b)
print(a%b)
print(divmod(a, b))

a=int(input())
b=int(input())
m=int(input())
print(pow(a,b))
print(pow(a,b,m))

a=int(input())
b=int(input())
c=int(input())
d=int(input())
k = a**b+c**d
print(k)

for i in range(1,int(input())):
    print(if"{i}"*)

n=int(input())
lst=set()
for i in range(n):
    country=input()
    lst.add(country)
print(len(lst))



# Djang s3 pip install django-storage
AWS_ACCESS_KEY_ID = ''
AWS_SECRET_ACCESS_KEY = ''
AWS_STORAGE_BUCKET_NAME = ''
AWS_S3_FILE_OVERWRITE = False
AWS_DEFAULT_ACL = None
DEFAULT_FILE_STORAGE = 'storages.backends.s3boto3.S3Boto3Storage'
STATICFILES_STORAGE = 'storages.backends.s3boto3.S3Boto3Storage'

string = input("Matn kiriting: ")
def swap_case(string):
    return string.swapcase()

result = swap_case(string)
print(result)

# test
class Test:
    def __init__(self, id1, id2):
        self.id1 = id1
        self.id2 = id2

    def test(self, other):
        if self.id1 == other:
            print("test-1: successfully")
        else:
            print("test-1: Wrong answer")

    def test2(self, other):
        if self.id2 == other:
            print("test-2: successfully")
        else:
            print("test-2: Wrong answer")

tst = Test(1, 2)
tst.test(1)
tst.test2(3)

# class, isinstance, TypeError, @classmethod and magic methods
class Clock:
    __DAY = 86400

    def __init__(self, seconds: int):
        if not isinstance(seconds, int):
            raise TypeError("Must be of type int")
        self.seconds = seconds % self.__DAY

    @classmethod
    def __verify_data(cls, other):
        if not isinstance(other, (int, Clock)):
            raise TypeError("Must be of type int")

        return other if isinstance(other, int) else other.seconds

    def __eq__(self, other):
        sc = self.__verify_data(other)
        return self.seconds == sc

    def __lt__(self, other):
        sc = self.__verify_data(other)
        return self.seconds < sc

    def __gt__(self, other):
        sc = self.__verify_data(other)
        return self.seconds > sc

    def __le__(self, other):
        sc = self.__verify_data(other)
        return self.seconds <= sc

    def __ge__(self, other):
        sc = self.__verify_data(other)
        return self.seconds >= sc

    def __ne__(self, other):
        sc = self.__verify_data(other)
        return self.seconds != sc

c1 = Clock(2000)
c2 = Clock(1000)
print(c1 == c2)
print(c1 < c2)
print(c1 > c2)
print(c1 <= c2)
print(c1 >= c2)
print(c1 != c2)

# hash object
class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
    
    def __eq__(self, other):
        return self.x == other.x and self.y == other.y

    def __hash__(self):
        return hash((self.x, self.y))

p1 = Point(1, 2)
p2 = Point(1, 2)

d = {}
d[p1] = 1
d[p2] = 2
print(d)
print(p1 == p2)
print(hash(p1) != hash(p2))
print(hash(p1), hash(p2), sep="\n")

from this import d
print(d)


# __len__ , __bool__
class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y

    def __len__(self):
        print("__len__")
        return self.x * self.x + self.y + self.y

    def __bool__(self):
        print("__bool__")
        return self.x == self.y

p = Point(10, 10)
# print(bool(p))
if p:
    print("p object gives True")
else:
    print("p object gives False")


    # working with class
from datetime import date
  
class Person:
    def __init__(self, name, age):
        self.name = name
        self.age = age
  
    # a class method to create a
    # Person object by birth year.
    @classmethod
    def fromBirthYear(cls, name, year):
        return cls(name, date.today().year - year)
  
    # a static method to check if a
    # Person is adult or not.
    @staticmethod
    def isAdult(age):
        return age > 18
    
    def get_age(self):
        return self.fromBirthYear("Erali", 22)
  
person1 = Person('mayank', 21)
person2 = Person.fromBirthYear('mayank', 1996)
  
print(person1.age)
print(person2.age)
print("Age:", person1.get_age())
  
# print the result
print(Person.isAdult(22))


class Meta(type):
    def create_local_attrs(self, *args, **kwargs):
        for key, value in self.class_attrs.items():
            self.__dict__[key] = value

    def __init__(cls, name, bases, attrs):
        cls.class_attrs = attrs
        cls.__init__ = Meta.create_local_attrs

class Women(metaclass=Meta):
    title = "Python"
    content = "content"
    photo = "image"

w = Women()
print(w.__dict__)


# method types
class MethodTypes:

    # class atributi
    name = "Ali"

    def instanceMethod(self):
        # self kalit so'zi orqali instance atributini yaratish
        self.lastname = "Valiyev"
        print(self.name, self.lastname, sep="\n")

    @classmethod
    def classMethod(cls):
        # cls kalit so'zi orqali class atributiga kirish
        cls.name = "Botir"
        print(cls.name)

    @staticmethod
    def staticMethod():
        print("Bu statik metod")


m = MethodTypes()
m.instanceMethod()

MethodTypes.classMethod()
MethodTypes.staticMethod()


#excel file to sqlite3 information import
import pandas as pd
import sqlite3

db = sqlite3.connect(':memory:')
dfs = pd.read_excel('BotBaza.xlsx', sheet_name=None)
for table, df in dfs.items():
    df.to_sql(table, db)

import sqlite3
import pandas as pd
filename="script"
con=sqlite3.connect("basebot.db")
wb=pd.ExcelFile('BotBaza.xlsx')
for sheet in wb.sheet_names:
        df=pd.read_excel(filename+'.xlsx',sheetname=sheet)
        df.to_sql(sheet,con, index=False,if_exists="replace")
con.commit()
con.close()


# class
class Geom:
    name = "Geom"

    def __init__(self, x1, y1, x2, y2):
        print(f"Geom ishlayapti! {self.__class__}")
        self.x1 = x1
        self.y1 = y1
        self.x2 = x2
        self.y2 = y2
        self.fill = 0


class Line(Geom):
    def draw(self):
        print("Chiziq chizish")

class Rect(Geom):
    def __init__(self, x1, y1, x2, y2, fill=None):
        super().__init__(x1, y1, x2, y2)
        print("Rect ishlayapti!")
        self.fill = fill

    def draw(self):
        print("Chiziq chizish")

l = Line(0, 0, 10, 20)
r = Rect(1,2,3,4)
print(r.__dict__)


# data science
from sklearn import tree

clf = tree.DecisionTreeClassifier()

# ma'lumotlarni yig'amiz
# quyoshli_kunlar, yomg'irli_kunlar, qorli_kunlar

x = [
    [220, 110, 35], [290, 60, 15],
    [260, 75, 30], [210, 140, 15],
    [200, 140, 25], [180, 180, 5],
    [180, 135, 50], [205, 120, 35],
    [110, 180, 75], [95, 197, 73]
]

y = [
    'yaxshi', 'yomon', "o'rtacha", 'yaxshi', "o'rtacha", 'yomon', 'yaxshi', "o'rtacha",        'yomon', 'yomon'
    ]

clf = clf.fit(x, y)
taxmin = clf.predict([[200, 65, 100]])
print(taxmin)


# magic methods
class Cat:
    def __init__(self, name):
        self.name = name

    def __repr__(self):
        return f"{self.__class__}: {self.name}"

    def __str__(self):
        return f"{self.name}"

    def __len__(self):
        return len(self.name)

    def __abs__(self, other):
        return other

c = Cat(name="Mosh")
print(str(c))
print(repr(c))
print(len(c))
print(abs(-5))


# all, map
def true_x(a):
    return a == 'x'

P = ['x', 'x', 'o', '0', 'x', 'o', 'x', 'x', 'x']

print(P[::3], P[1::3], P[2::3])

row_1 = all(map(true_x, P[:3]))
row_2 = all(map(true_x, P[3:6]))
row_3 = all(map(true_x, P[6:]))

col_1 = all(map(true_x, P[::3]))
col_2 = all(map(true_x, P[1::3]))
col_3 = all(map(true_x, P[2::3]))

print(row_1, row_2, row_3)
print(col_1, col_2, col_3)


# matndan ovozga
# import pyttsx3
# import os
# dis2 = pyttsx3.init()
# dis2.setProperty('rate', 150)
# dis2.setProperty('volume', 0.7)
# fil=open('aql.txt', 'r')
# matn=fil.read()
# # matn = input("Matn kiriting: ")
# dis2.say("Assalomu alaykum! Pythonda TTS dan foydalanish.")
# dis2.say(matn)
# dis2.save_to_file(matn,'test.mp3')
# dis2.runAndWait()
# os.system('music/test.mp3')

# class Person:
#     def __init__(self, first_name, last_name, age):
#         self.first_name = first_name
#         self.last_name = last_name
#         self.age = age

#     def info(self):
#         return f"Name: {self.first_name}\nSurname: {self.last_name}\nAge: {self.age}"

# class Student(Person):
#     def __init__(self, first_name, last_name, age, group):
#         super().__init__(first_name, last_name, age)
#         self.group = group

#     def info(self):
#         return f"Name: {self.first_name}\nSurname: {self.last_name}\nAge: {self.age}\nGroup: {self.group}"

# student = Student("Ali", "Olimov", 23, "9-19")
# print(student.info())

# n = int(input("Enter the number: "))
# suma = 0
# mult = 1
 
# while n > 0:
#     digit = n % 10
#     suma = suma + digit
#     mult = mult * digit
#     n = n // 10
 
# print("Total:", suma)
# print("Multiplication:", mult)


# from random import randint
# from math import log10, floor

# def human_format(num, ends=["", "K", "M", "B", "T"]):
#     # divides by 3 to separate into thousands (...000)
# 	return ends[int(floor(log10(num))/3)]

# if __name__ == '__main__':
#     for i in range(10):
#         x = randint(1,10**i)
#         print(x, human_format(x))


####################################################

# from collections import Counter, defaultdict

# def build_counter(word):
#     counter = defaultdict(int)
#     for char in word:
#         counter[char] += 1
#     return counter

# def is_anogram(word1:str, word2:str) -> bool:

#     counter1 = Counter(word1)
#     counter2 = Counter(word2)
#     print(counter1)
#     print(counter2)
#     return counter1 == counter2


# print(is_anogram("mosh", "shom"))

################################################

# from collections import Counter, defaultdict

# def build_counter(word):
#     counter = defaultdict(int)
#     for char in word:
#         counter[char] += 1
#     return counter

# def is_anogram(word1:str, word2:str) -> bool:

#     counter = build_counter(word1)

#     for char in word2:
#         counter[char] -= 1
#     return all(val == 0 for val in counter.values())
#     # for val in counter.values():
#     #     if val != 0:
#     #         return False
#     # return True

# print(is_anogram("mosh", "shom"))

###############################################
# def build_counter(word):
#     counter = dict()
#     for char in word:
#         if char not in counter:
#             counter[char] = 1
#             continue
#         counter[char] += 1
#     return counter


# def is_anogram(word1:str, word2:str) -> bool:

#     counter1 = build_counter(word1)
#     counter2 = build_counter(word2)
#     print(counter1)
#     print(counter2)
#     return counter1 == counter2


# print(is_anogram("mosh", "shom"))
######################################################


#######################################################
# def is_anogram(word1:str, word2:str) -> bool:

#     counter1 = dict()
#     counter2 = dict()

#     for char in word1:
#         if char not in counter1:
#             counter1[char] = 1
#             continue
#         counter1[char] += 1

#     for char in word2:
#         if char not in counter2:
#             counter2[char] = 1
#             continue
#         counter1[char] += 1

#     return counter1 == counter2

# print(is_anogram("mosh", "shom"))

#####################################################


# Misol: 
# two_sum([1,4,5,2], 3) => True
# two_sum([1,4,5,2], 8) => False


# def two_sum(arr: list[int], target: int) -> bool:
#     complements = {}

#     for index, num in enumerate(arr):
#         print("ayr:",target - num)
#         complements[target - num] = index
#     print(complements)

#     for index, num in enumerate(arr):
#         if num in complements and complements[num] != index:
#             # print(num, index)
#             return True
#     return False

# print(two_sum([1,4,5,2], 3))


# def enumerate_func(lst: list[int]) -> int:

#     for index, num in enumerate(lst):
#         print('index:',index, 'number:',num)
# print("Index va o'sha index'dagi son:\n")
# enumerate_func(lst = [12,25,3,4,42])


# regex
import re
from tkinter import N
# re.search
# boshlanishi va tugash va boshqa hallarni tekshiradi.
text = "hello guys what is guys"

text_start = re.search('^h', text)
text_end = re.search('s$', text)
text_none = re.search('y$', text)

print(text_start)
print(text_end)
print(text_none)

if text_start and text_end:
    print('Yes')
else:
    print('No')

# re.findall
# matnni ichidan kiritgan so'zni izlab topib list ko'rinishida qirqib oladi.
re_text1 = re.findall('^hello', text)
print(re_text1)

# re.split
# bo'sh joy bo'yicha split qiladi.
text_split = re.split("\s", text)
text_split_2 = re.split("\s", text, 2)
print(text_split)
print(text_split_2)

# re.sub
# bo'sh joyni berilgan belgi bilan almashtirib beradi va soni bo'yicha ham yuqoridagi ishni bajaradi
text_sub = re.sub("\s", "5", text)
text_sub_2 = re.sub("\s", "5", text, 2)
print(text_sub)
print(text_sub_2)




n = 125

S = n ** 2
P = 4 * n

print("Yuzasi:",S)
print("Peremetri:",P)


####### 2-qism ###########
# regex
import re
# re.search
# boshlanishi va tugash va boshqa hallarni tekshiradi.
text = "hello guys what is hi"

text1= re.search(r'\bg\w+', text)
text2 = re.search(r'\bg\w+', text)
text3 = re.search(r'\bg\w+', text)


# span() boshlanish va yakuniy pozitsiyalarini o'z ichiga olgan tuple qaytaradi.
print(text1.span())
# funktsiyaga kiritilgan satrni chop etadi.
print(text2.string)
# group() satrning mos keladigan qismini qaytaradi
print(text3.group())


######## dict and input #########
# user_info = dict()

# name = input(">>> Enter Name:")
# user_info["name"] = name

# username = input(f">>> {name} enter username:")
# user_info["username"] = username

# email = input(f">>> {name} enter email:")
# user_info["email"] = email

# password = input(f">>> {name} enter password:")
# user_info["password"] = password

# print(f"User info:\n\nName: {name}\nUsername: {username}\nE-mail: {email}\nPassword: {password}")
# print(f"Dictionary:\n{user_info}")

####### list ##########
mylist = ["apple", "banana", "cherry", "orange", "kiwi", "melon", "mango"]
print(mylist)
# ro'yxat elementlariga kirish
print(mylist[0])
print(mylist[-1])
print(mylist[:2])
print(mylist[4:])
print(mylist[-3:-1])
print(mylist[2:6])

# ro'yxat elementlarini o'zgartirish
mylist[1] = "watermelon"
print(mylist)
mylist[1:2] = ["blackcurrant", "watermelon"]
print(mylist)
# ro'yxatga element qo'shish
# insert() - berilgan indexga ro'yxat elementini kiritish
mylist.insert(4, "watermelon")
print(mylist)
mylist.append("watermelon")
print(mylist)
tuple_e = ("kiwi", "orange")
mylist.extend(tuple_e)
print(mylist)

# # rjust
# string = "hello everybody"
# right_just = string.rjust(len(string) + 10, '*')
# print(right_just)
# right_just1 = string.rjust(len(string) + 10, '9')
# print(right_just1)
# right_just2 = string.rjust(len(string) + 10, 'a')
# print(right_just2)

# # ljust
# string = "hello everybody"
# left_just = string.ljust(len(string) + 10, '*')
# print(left_just)
# left_just1 = string.ljust(len(string) + 10, 'a')
# print(left_just1)
# left_just2 = string.ljust(len(string) + 10, '5')
# print(left_just2)

# random sorting
from random import randint

n = int(input("Enter a Number: "))
number_list = []
for i in range(n):
    number_list.append(randint(1, 99))
print(number_list)

for i in range(n - 1):
    for j in range(n-i-1):
        if number_list[j] > number_list[j+1]:
            number_list[j], number_list[j+1] = number_list[j+1], number_list[j]
print(number_list)


# class
class Car:
    def __init__(self, top_speed: int) -> None:
        self.top_speed = top_speed

    def drive(self):
        print("Driving")

def create_fast_car_(car_class: type[Car], top_speed: int):
    return car_class(top_speed=top_speed)
create_fast_car_(Car, 120).drive()


import numpy as np
import scipy as sp
import matplotlib as mp
import pandas as pd
import sklearn as sl
import seaborn as sb

import numpy
import scipy
import matplotlib
import pandas
import sklearn
import seaborn

# pip install numpy
# pip install pandas
# pip install scipy
# pip install matplotlib
# pip install seaborn
# pip install scikit-learn

arr = np.array([1, 2, 3, 4], dtype=int)

print(arr)
print(type(arr))

# string format
number = 50
text = "Number: {}"
text1 = "Number: {:.4f}"
print(text.format(number))
print(text1.format(number))

years = 7
months = 254
days = 52
my_order = "Years: {} Months: {} Days {:.2f}"
my_order1 = "Years: {0} Months: {1} Days {2:.2f}"
print(my_order.format(years, months, days))
print(my_order1.format(years, months, days))

my_order2 = "Car Name: {car_name}, Model: {model}."
print(my_order2.format(car_name = "Malibu", model = "GM"))

age = 36
name = "Olim"
info = "His name is {1}. {1} is {0} years old."
print(info.format(age, name))


# "%s" string format operatori
print("Hello guys, %s is format"%"this")

text = "you"
print("What do %s do?"%(text))
print("Are %s learning %s?"%(text, 'English'))

# butun sonlar uchun '%d' format operatori ishlatiladi.
print("John %d years old"%15)

# %a.bf
# a satrda mavjud bo'lgan raqamlarning minimal soni hisoblanadi.
# "bf" kasrdan keyin qancha raqam ko'rsatilishini bildiradi.
print('The value of pi is: %3.2f'%(3.141592))
print('Floating point numbers: %1.0f' %(13.144))

variable = 12
string = "Variable as integer = %d \n\
Variable as float = %f" %(variable, variable)
print (string)


# f-string format
text = 'Format'
print(f"{text}, this is {text}.")
 
 
name = 'Alisher'
age = 21
print(f"Hello, My name is {name} and I'm {age} years old.")

import datetime
# f-string vaqtlar bilan ishlashga qulay
# bu ro'yxatni yana kengaytirish mumkin
today = datetime.datetime.today()
print(f"{today:%B %d, %Y}")
print(f"{today:%d-%m-%Y}")
print(f"{today:%D}")
# F-string eng ko'p qo'llaniladigan ikkita string formatlash mexanizmlaridan tezroqdir, ular % formatlash va str.format().

###########################################
Python-da Array va List farqi 👇🏻

Ro'yxat(List) Pythonga oʻrnatilgan va elementlar toʻplamini saqlaydigan maʼlumotlar strukturasidir. 
Ro'yxat xususiyatlari: tartiblangan, oʻzgaruvchan, obyektlar takrorlanadi, har xil turdagi ma'lumotlarga ega bo’ladi.

Massiv(Array) esa elementlar to'plamini saqlaydigan ma'lumotlar strukturasidir. 
Massiv xususiyatlari: tartiblangan, oʻzgaruvchan, obyektlar takrorlanadi.

Ammo massivning turli xil ma'lumotlar turlarini saqlash qobiliyati haqida gap ketganda, javob unchalik oddiy emas. Bu ishlatiladigan massiv turiga bog'liq.

Python-da massivlardan foydalanish uchun siz massiv modulini yoki NumPy paketini import qilishingiz kerak .

import array as arr
import numpy as np

array moduli barcha massiv elementlarini bir xil turdagi bo'lishini talab qiladi. Va qiymat turini kiritishingiz kerak bo’ladi.

array1 = arr.array("i", [2, 4, 8, 16])

Boshqa tomondan, NumPy massivlari turli xil ma'lumotlar turlarini qo'llab-quvvatlaydi.

array2 = np.array(["numbers", 2, 4, 8, 16])

Xo’sh farq nimada? 🤔

Massivlar e'lon qilinishi kerak. Ro'yxatlar yo'q. 
Massivlar ma'lumotlarni juda ixcham saqlashi mumkin va katta hajmdagi ma'lumotlarni saqlash uchun samaraliroqdir.
Massivlar raqamli operatsiyalar uchun juda yaxshi; ro'yxatlar to'g'ridan-to'g'ri matematik operatsiyalarni bajara olmaydi. Masalan, massivning har bir elementini faqat bitta kod qatori bilan bir xil raqamga bo'lishingiz mumkin. Agar siz ro'yxat elementini bilan bir xil raqamga bo’lsangiz, xatoga yo'l qo'yasiz.

array = np.array([3, 6, 9, 12])
division = array/3
print(division)
# [1. 2. 3. 4.]

lst = [3, 6, 9, 12]
division = lst/3
print(division)
# TypeError: unsupported operand type(s) for /: 'list' and 'int'

Albatta, ro‘yxat yordamida matematik amalni bajarish mumkin, ammo bu unchalik samarali emas.
#############################################

# callback function
def callback(a, b):
    print(f'Sum = {a+b}')

def main(a,b,func=None):
    print('Add any two numbers')
    if func is not None:
        func(a,b)

one, two = map(int, input("Enter numbers: ").split())
main(one, two, callback)


# Speach to text Stt
# __author__ = 'NORMATOV S.'
# import speech_recognition as stt
# r = stt.Recognizer()
# with stt.Microphone() as source:
#     print("Gapirishni boshlang, vaqt ketdi");
#     audio = r.listen(source)
#     print("Vaqt tugadi. Rahmat")
# try:
#     print("Text: " + r.recognize_google(audio));
# except:
#     pass;
    
# Python  Dictionary
person = {
  "first_name": "Alisher",
  "last_name": "Aliyev",
  "age": 22,
  "email": "alisheraliyev@gmail.com",
  "username": "aliali",
}

items = person.items()
keys = person.keys()
values = person.values()

print(f"Items: {items}\nKeys: {keys}\nValues: {values}")


# OpenPyXL
# pip install openpyxl
import openpyxl

def user_data_read(file_name):
    users = openpyxl.open(file_name, read_only=True)
    sheet = users.active
    # print(users)

    data = []
    keys = ['user_id', 'first_name', 'last_name', 'age']

    for row in sheet.iter_rows(min_col=2, max_col=5, min_row=2, max_row=5):
        data_dict = dict(zip(keys, (cell.value for cell in row)))
        data.append(data_dict)
        # print(data_dict)
    print(data)
user_data_read("user_data.xlsx")


# Python-da polimorfizm-ning bir ko'rinishi
class Error:
    def get_result(self):
        return "get_result() method does not exist"
    
class Restangle(Error):
    def __init__(self, w, h) -> None:
        self.w = w
        self.h = h

    def get_result(self):
        return 2*(self.w + self.h)
    

class Square(Error):
    def __init__(self, n) -> None:
        self.n = n

    def get_result(self):
        return 4*self.n

class Triangle(Error):
    def __init__(self, a, b, c) -> None:
        self.a = a
        self.b = b
        self.c = c
    
    # def get_result(self):
    #     return self.a + self.b + self.c
    
r1 = Restangle(3, 4)
r2 = Restangle(5, 6)
s1 = Square(5)
s2 = Square(8)
t1 = Triangle(3, 4, 5)
t2 = Triangle(1, 3, 6)

lst = [r1, r2, s1, s2, t1, t2]
for l in lst:
    print(l.get_result())

####################
# year = int(input("Yilni kiriting: "))

# if year == 2022:
#     print(f"Siz {year}-yilning birinchi qish kunidasiz!")
# elif year < 2022 and len(str(year)) >=4:
#     print("Adashdingiz. Mantiqan xato!")
# elif year > 2022:
#     print(f"Siz {year - 2022} yilga qariysiz!")
# else:
#     print("Don't stupid")
###################

# break
text = "python"

# for
for letter in text:
    print(letter)
    if letter == "t":
        break
print("For tsiklidan tashqarida!")
print()

# while
i = 0
while True:
    print(text[i])

    if text[i] == 't':
        break
    
    i += 1
print("While tsiklidan tashqarida!")

# continue
for i in range(1, 10):

    if i == 4 or i == 7:
        continue
    else:
        print(i, end=" ")

# pass
print()
text = "cup"
for t in text:

    if t == "u":
        print("pass bajarildi!")
        pass
    print(t)

   
#############
# 3 ta sondan kattasini topish dasturi
import time
a, b, c = list(map(int, input("Sonlarni kiriting: ").split()))

start1 = time.time()
if a > b and a > c:
    print(f"{a} soni eng katta")
elif b > a and b > c:
    print(f"{b} soni eng katta")
elif c > a and c > b:
    print(f"{c} soni eng katta")
end1 = time.time()

start2 = time.time()
if a > b:
    if a > c:
        print(f"{a} soni eng katta")
    else:
        print(f"{c} soni eng katta")
else:
    if b > c:
        print(f"{b} soni eng katta")
    else:
        print(f"{c} soni eng katta")
end2 = time.time()

time1 = end1 - start1
time2 = end2 - start2
print(time1, time2)


# Python-da statistics moduli
# o'rtacha qiymatni hisoblash mean() funksiyasi orqali
import statistics

lst1 = [1, 3, 5, 7, 9, 11, 13]
lst2 = [1, 3, 5, 7, 9, 11]

avg1 = statistics.mean(lst1)
avg2 = statistics.mean(lst2)

print(f"Average1: {avg1}\nAverage2: {avg2}")

############### filter() ############
def func(variable):
    letters = ['a', 'i', 'e', 'o', 'u']
    if variable in letters:
        return True
    else:
        return False

sequence = ['g', 'e', 'm', 'j', 'u', 'k', 's', 'p']
filtered = filter(func, sequence)

print("Filtrlangan harflar:")
for f in filtered:
    print(f)
# 2-usul
letters = ['a', 'i', 'e', 'o', 'u']
sequence = input("Harflarni kiriting: ")

check = list(filter(lambda x : x in letters, sequence))
print(f"Filtrlangan harflar:\n{check}")


####### linking classes ###########
# class-larning bir-biri bilan bog'lanishi
class Salary:
    def __init__(self, pay):
        self.pay = pay

    def get_total(self):
        return self.pay * 12

class Employee:
    def __init__(self, pay, bonus):
        self.pay = pay
        self.bonus = bonus
        self.salary = Salary(self.pay)

    def get_bonus(self):
        return self.bonus * 12
    
    def get_salary(self):
        return f"Total salary: ${self.salary.get_total() + self.get_bonus()}"

employee = Employee(500, 200)
print(employee.get_salary())


############################################
a = int(input("a sonini kiriting: "))
b = int(input("b sonini kiriting: "))

if a < b:
    for i in range(a, b+1):
        for j in range(i):
            print(i, end=" ")
        print("")
else:
    print("Xato! a soni b sonidan katta va teng bo'lmasligi kerak.\n"
        "Eslatma: a soni b sonidan kichik bo'lishi shart!")
    
############################################

from translate import Translator

res_lst = ['Natija', 'Result', 'Результат']
from_lans = input("Tilni kiriting: ")
word = input("Matnni kiriting: ")
to_lans = input("Tilni kiriting: ")

translator = Translator(from_lang=from_lans, to_lang=to_lans)
translation = translator.translate(word)

print(translation)

########## FastAPI #########
from typing import Union

from fastapi import FastAPI

app = FastAPI()


@app.get("/")
def read_root():
    return {"result": "Hello World"}


@app.get("/items/{item_id}")
def read_item(item_id: int, query: Union[str, None] = None):
    return {"item_id": item_id, "query": query}

######### Web Server #############
# Python yordamida kichik web server ning ko'rinishi
from http.server import HTTPServer, BaseHTTPRequestHandler

class web_server(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/':
            self.path = '/index.html'
        try:
            file_to_open = open(self.path[1:]).read()
            self.send_response(200)
        except:
            file_to_open = "Not Found Error:404"
            self.send_response(404)
        self.end_headers()
        self.wfile.write(bytes(file_to_open, 'utf-8'))
httpd = HTTPServer(('localhost', 8080), web_server)
httpd.serve_forever()

########## Django models ###########
from django.db import models
from django.contrib.auth.models import User

class Post(models.Model):
    # relations
    admin = models.ForeignKey(User, on_delete=models.SET_NULL, null=True, blank=True)
    # fields
    image = models.ImageField(upload_to='post_img')
    title = models.CharField(max_length=255, null=True, blank=True)
    body = models.TextField()
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

 
    
####### match case ########
def http_error(status):
    match status:
        case 400:
            return "Bad request"
        case 404:
            return "Not found"
        case 418:
            return "I'm a teapot"
        case _:
            return "Something's wrong with the internet"
status = int(input("Enter a status: "))
print(http_error(status=status))

######## Dict(Lug'at) elementlariga kirish ########
dictionary = {
    "person1": {
        "name": "John Doe",
        "age": 24,
        "job": "Doctor",
        "car": "Mustang"
    },

    "person2": {
        "name": "Jek Johns",
        "age": 32,
        "job": "Programmer",
        "car": "Tesla"
    },

    "skills": [
        "Python", "Java", "C++", "PHP", "GO", "C#", "JavaScript"
    ],
}
name = dictionary["person1"]["name"]
print(name)

# job = dictionary.get("person2").get("job")
job = dictionary.get("person2")["job"]
print(job)

skill = dictionary.get("skills")[0]
print(skill)


####### Paskal uchburchagi ######
# Paskal uchburchagi
n = int(input("n sonini kiriting: "))
print("Paskal uchburchagi:")

for i in range(1, n+1):
    for j in range(0, n-i+1):
        print(' ', end='')

    c = 1
    for j in range(1, i+1):
        print(' ', c, sep='', end='')
        c = c * (i - j) // j
    print()

    
#######  IP Address  #######    
# # IP Address
# import ipaddress
# net = ipaddress.ip_network('74.125.227.0/29')

# for  address in net:
#     print(address)    

###### Video to Gif ######
# Video to Gif
from moviepy.editor import VideoFileClip
clip = VideoFileClip("video.mp4")
clip.write_gif("mygif.gif")

####### Write text on video  ########
from moviepy.editor import VideoFileClip, TextClip, CompositeVideoClip

video = VideoFileClip("talaba.mp4").subclip(50,60)
txt_clip = (TextClip("Talabalik davrim 2023", fontsize=70, color='black')
            .set_position('center')
            .set_duration(10) )

result = CompositeVideoClip([video, txt_clip])
result.write_videofile("my_student_period.webm",fps=25)

######## Connection to PostgreSQL in Python  #########
import psycopg2
from pprint import pprint

conn = psycopg2.connect(
   database="lesson",
   user='postgres', 
   password='aeb1205sql', 
   host='localhost', 
   port= '5432'
)
cursor = conn.cursor()

cursor.execute("""SELECT * FROM users ORDER BY id""")
pprint(cursor.fetchall())

pprint("CONNECTED")

#############  Anagram check   #############
# Anagrm tekshiruvi
from collections import Counter

str1 = 'Python'
str2 = 'thyPon'

print(Counter(str1) == Counter(str2))

str3 = 'Python love'
str4 = 'Python hate'

print(Counter(str3) == Counter(str4))

# Email slicer with Python
email = input("Enter Your Email: ").strip()

username = email[:email.index("@")]
domain_name = email[email.index("@"):]

result = f"Your user name is '{username}' and your domain is '{domain_name}'"
print(result)

# shuffle function
from random import shuffle

num_lst = [1, 2, 3, 4, 5, 6, 7, 8, 9]
letter_lst = ['A', 'B', 'C', 'D', 'E', 'F']

shuffle(num_lst)
shuffle(letter_lst)

print(num_lst)
print(letter_lst)


# Send an SMS Using Twilio

  import os
  from twilio.rest import Client


  account_sid = "ACba1cf469702bb254c4aadc4ff437e930"  
  auth_token  = "your_auth_token"
  #account_sid = os.environ['TWILIO_ACCOUNT_SID']
  #auth_token = os.environ['TWILIO_AUTH_TOKEN']

  client = Client(account_sid, auth_token)

  message = client.messages.create(
      from_="+99899XXXXXXX",
      to="+99899XXXXXXX",
      body="Hello my friend")

  print(message.sid)

    
# # Online shop 1-method
# n = int(input("Nechta mahsulot olasiz sonini kiriting: "))
# mahsulotlar = ["tuz", "shakar", "un", "yog'", "guruch", "go'sht", "non", "qaymoq", "gugurt", "piyoz"]

# savat = []
# i = 1
# while i > 0:
#     if i == n+1:
#         break
#     kirit = input(f"{i}-mahsulotni kiriting: >>>")
#     if kirit in mahsulotlar:
#         savat.append(kirit)
#         print("Mahsulot do'konimizda bor!")
#     else:
#         print("Mahsulot do'konimizda yo'q!")
    
#     i += 1
# print(f"Bizning do'kondan xarid qilgan mahsulotlaringiz:\n{savat}")

# Online shop 2-method
n = int(input("Nechta mahsulot olasiz sonini kiriting: "))
mahsulotlar = ["tuz", "shakar", "un", "yog'", "guruch", "go'sht", "non", "qaymoq", "gugurt", "piyoz"]

savat = []

for i in range(n):
    kirit = input(f"{i+1}-mahsulotni kiriting: >>> ")

    if kirit in mahsulotlar:
        savat.append(kirit)
        print("Mahsulot do'konimizda bor!")
    else:
        print("Mahsulot do'konimizda yo'q!")

print(f"Bizning do'kondan xarid qilgan mahsulotlaringiz:\n{savat}")


# # Palindrom number in python
# num = int(input("Enter a value:"))

# def is_palindrom(num): 
#     temp = num  
#     rev = 0  
#     while(num > 0):  
#         dig = num % 10  
#         rev = rev * 10 + dig
#         print(rev)
#         num = num // 10
#     if(temp == rev):
#         return True 
#     else:  
#         return False
# print(is_palindrom(num=num))


def searchInsert(nums: List(int), target: int) -> int:
    l = len(nums)
    if target in nums:
        print("-- working -- 1")
        return nums.index(target)
    if target not in nums:
        nums.append(target)
        sorted_nums = sorted(nums)
        print("-- working -- 2")
        return sorted_nums.index(target)
print(searchInsert([-1,3,5,6], 0))


# class Class1:
#     def func1(self):
#         print("func1() method of class Class1")

# class Class2(Class1):
#     def func2(self):
#         print("func2() method of class Class2")

# class Class3(Class1):
#     def func1(self):
#         print("func1() method of class Class3")
#     def func2(self):
#         print("func2() method of class Class3")
#     def func3(self):
#         print("func3() method of class Class3")
#     def func4(self):
#         print("func4() method of class Class3")

# class Class4(Class2, Class3):
#     def func4(self):
#         print("func4() method of class Class4")

# c1 = Class4()
# c1.func1()
# c1.func2()
# c1.func3()
# c1.func4()

# import pyautogui
# screenshot = pyautogui.screenshot()
# screenshot.save("screenshot.png")
# import pyautogui
# pyautogui.alert('This is an alert box.')
    
# writing unittest to Python OOP
class First:
    def first(self):
        print("first() method of class First")

class Two:
    def two(self):
        print("two() method of class Two")

class Three(First, Two):
    pass
 
a = Three()

if isinstance(a, First) and isinstance(a, Two) and isinstance(a, Three):
    print("Ran 3 tests in 0.001s\n\nOK")
else:
    print("Test failed!")


import unittest

class TestStringMethods(unittest.TestCase):

    def test_first(self):
        self.assertTrue(a, First)

    def test_two(self):
        self.assertTrue(isinstance(a, Two))

    def test_split(self):
        self.assertIsInstance(a, Three)

if __name__ == '__main__':
    unittest.main()
    
###################################
# def myPow(x: float, n: int) -> float:
#     result = f"{x**n:.5f}"
#     return float(result)
    

# print(myPow(2.00000, 2))

#set remove and discard
thisset = {"apple", "banana", "cherry"}

thisset.remove("banana")
# thisset.discard("banana")

print(thisset)



from fpdf import FPDF
import uuid
pdf_name = str(uuid.uuid4()).split('-')[-1]
# print(pdf_name)
pdf = FPDF('P', 'mm', 'A4')

pdf.add_page()

pdf.image("pyy.png", 50, 50, 100)

pdf.set_font("Arial", 'B', size=14)
pdf.set_text_color(220, 50, 50)

pdf.cell(200, 10, txt="Theme: About FPDF class", ln=1, align='C')

f = open('D:\Python\python_OOP/aql.txt', "r")
for i in f:
    pdf.cell(200, 10, txt = i, ln = 1, align = 'C')

pdf.cell(200, 10, txt = "Write to Center", ln = 2, align = 'C')
pdf.cell(200, 10, txt = "Write to Left", ln = 3, align = 'L')
pdf.cell(190, 10, txt = "Write to Right", ln = 3, align = 'R')

pdf.output(f"pdf_{pdf_name}.pdf")

# file open with python
file = open("aql.txt", "r")
f = file.read()
print(f)

file.close()

with open('aql.txt', 'r') as file:
    f = file.read()
    print(f)
    
    
#########################################
# setattr in Python
class Person:
    name = "John"
    age = 35

person = Person()

setattr(person, "job", "programmer")
print(person.job)

setattr(person, "name", "Jek")
print(person.name)

# getattr in Python
class Person:
    name = "John"
    age = 35

person = Person()

name = getattr(person, "name")
print(name)

job = getattr(person, "job", "programmer")
print(job)

try:
    job_err = getattr(person, "job")
    print(job_err)
except AttributeError:
    print("Attribute is not found!")

    
############################
simple example
def in_range(*args):
    if args[2] >= args[0] and args[2] <= args[1]:
        res = f"Berilgan {args[2]} soni {args[0]} va {args[1]} sonlari orasida"
    else:
        res = f"Berilgan {args[2]} soni {args[0]} va {args[1]} sonlari orasida emas !!!"
        
    return res

minimum, maximum, give_number = map(int, input("Enter numbers: ").split())
print(in_range(minimum, maximum, give_number))

# calculator
def simple_calculator(num1, num2, operator):
    match operator:
        case "add":
            print(f"Sum: {num1 + num2}")
        case "subtract":
            print(f"The difference: {num1 - num2}")
        case "multiply":
            print(f"Multiple: {num1 * num2}")
        case "divide":
            print(f"Division: {num1 / num2}")
        case _:
            print("Undefined action")

num1, num2 = map(int, input("Enter numbers: ").split())
operator = input("Enter operator (for example: 'add', 'subtract', 'multiply', 'divide') >>> ")
simple_calculator(num1, num2, operator)

# largest
def get_largest_element(nums: List(int)):
    largest = nums[0]
    for i in range(len(nums)):
        if nums[i] > largest:
            largest = nums[i]

    return largest

print(get_largest_element([4, 55, 9, 23, 98, 11, 16, 19, 37, 745]))

########################
have and no
input_word = input("O'ylagan so'zingizni kiriting: ")
word = bytes(input_word, encoding="utf-8")

with open("stop_words(uz).txt", "rb") as f:
    words = f.read()

    print(words)

    str_value = words.decode("utf-8")
    list_value = str_value.split()
 
    if word in words:
        print("True")
    else:
        print("False")
        
 ###########

# Solution
def roman_to_int(s):
    roman_dict = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}

    result = 0

    for i in range(len(s)):
        if i > 0 and roman_dict[s[i]] > roman_dict[s[i - 1]]:
            print("1: ", roman_dict[s[i]] - 2 * roman_dict[s[i - 1]])
            result += roman_dict[s[i]] - 2 * roman_dict[s[i - 1]]
        else:
            print("2: ", roman_dict[s[i]])
            result += roman_dict[s[i]]
    return result

input_roman = input("Enter Roman a number: ")
print(roman_to_int(input_roman))

def twoSum(nums, target):
    nums_dict = {}
    for i, num in enumerate(nums):
        complement = target - num
        if complement in nums_dict:
            return [nums_dict[complement], i]
        nums_dict[num] = i
    return []
print(twoSum([2,7,11,15],  9))


# delete right and left space of the text
text = "   Python Dev   "

# text length
print(len(text))
# 16

# print(len(text.lstrip()))
# print(len(text.rstrip()))

# delete empty place with strip() method
clean_text = text.strip()

print(clean_text)
# Python Dev

# clean text length
print(len(clean_text))
# 10


# Django: Slug
from django.db import models
from django.urls import reverse

class Post(models.Model):
    title = models.CharField(max_length=255)
    body = models.TextField()
    slug = models.SlugField()

    def __str__(self):
        return self.title
    
    def get_absolute_url(self):
        return reverse("post_detail", kwargs={"slug": self.slug})
    
    
####################
# walrus operator in Python
words = []

while (word := input("Enter word: ")) != "quit":
    words.append(word)

print()
print("*"*33)
print(words)

n = (m := 5)
print(n == m)


# schedule library in Python
import schedule
import time

def task1():
    print("Running task 1...")

def task2():
    print("Running task 2...")

def task3():
    print("Running task 3...")

# task1, task2 va task3 funksiyalarini 
# har daqiqada, sekundda va soatlarda o'z vazifasini bajarishini rejalashtiring
schedule.every().seconds.do(task1)
schedule.every(1).minutes.do(task2)
schedule.every(1).hours.do(task3)

# task2 func har kuni soat 07:00 da ishga tushadi.
schedule.every().day.at("07:00").do(task2)

# task3 func har dushanba kuni 12:00 da ishga tushadi.
schedule.every().monday.at("12:00").do(task3)

while True:
    schedule.run_pending()
    time.sleep(1)
    
####################
n = int(input("Enter a number: "))

for i in range(1, n+1):
    left_space = ". " * (n-i)
    right_space = ". " * (n-i)
    series = ("0 ") * i + "0 " * (i-1)
    print(left_space + series + right_space)
    
    
#####################
# Dictionary get() method in Python
# Sintaksisi: Dict.get(key, default=None)

d = {'coding': 'good', 'thinking': 'better'}
print(d.get('coding'))

d = {1: '001', 2: '010', 3: '011'}
# agar 4 kalit so'zi topilmasa default qiymat qaytariladi(Not found)
print(d.get(4, "Not found"))

test_dict = {'coding': {'is': 'best'}}

# original dictionary
print("Dictionary: " + str(test_dict))

# dictga ichma-ich kirish
res = test_dict.get('coding', {}).get('is')

# result
print("Value: " + str(res))

#########################
Python Encapsulation
class MyList:
    def __init__(self, lst):
        self.__lst = lst

    def my_func(self):
        res = self.__lst[0][2] + sum(self.__lst[1][1:])
        return res
    
    def lst(self):
        return self.__lst

my_list = MyList([[1, 2, 5], [6, 8, 10]])
print(my_list.my_func())

######################
# boltons library in Python

from boltons import strutils

text = "Hello, World"

print(strutils.split_punct_ws(text))
# ['Hello', 'World']

print(strutils.html2text('<div>Hello friends!</div>'))
# Hello friends!

print(strutils.find_hashtags('Yes #pip Okey #install Hi #boltons'))
# ['pip', 'install', 'boltons']

print(strutils.slugify(text))
# hello_world

#################
# import json

# with open('data_file.json', 'r') as file:
#     json_file = json.load(file)

# print(json_file)
# # print(type(json_file))

# output_value_name = json_file['name']
# print(output_value_name)

# outpu_value = json_file['programming_language']['Python']
# print(outpu_value)

# dict_file = {"hi": "salom"}
# json_dict_file = json.dumps(dict_file)
# print(type(json_dict_file))

#####################
# Creating a decorator function in Python
def validate_inputs(func):

    def wrapper(*args, **kwargs):
        for arg in args:
            if not isinstance(arg, int):
                raise TypeError("All arguments must be integers")
        for value in kwargs.values():
            if not isinstance(value, int):
                raise TypeError("All keyword arguments must be integers")

        return func(*args, **kwargs)

    return wrapper


@validate_inputs
def calculate_average(*args, **kwargs):
    total_sum = sum(args) + sum(kwargs.values())
    count = len(args) + len(kwargs)
    return total_sum / count


print(calculate_average(1, 2, 3, x=4, y=5, z=6))
print(calculate_average(1, "two", x=3, y=4))

##############
#time() module in Python

import time

# time() - 1970-yil 1-yanvar 00:00:00 (UTC) dan 
# beri o‘tgan soniyalar sonini qaytaradi.
current_time = time.time()
print("Current time:", current_time)

# soniyalarni yilga o'tkazish
until_now_year = ((current_time/3600)/24)/365
print("Until now:", until_now_year)

# localtime() - "Sun Jun 20 23:21:05 1993” formatidagi qatorga aylantiradi.
local_time = time.localtime()
print("Local time:", local_time)

# strptime(string, format)- Format satrida belgilangan formatdagi 
# vaqtni ifodalovchi qatorni tahlil qiladi va struct_time ni qaytaradi.

# strftime() yordamida joriy vaqtni satr sifatida formatlang
strf_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
print("Time:", strf_time)

#####################
# Example to dict and function
def home():
    menu = input("""
            users list:       1
            add user:         2
            update user info: 3
            delete user:      4
            search user:      5
            quit:             0
        """)
    if menu.isdigit():
        main(menu)
    else:
        print("Command must be number")
        home()

users = {}

def add_user():
    name = input("Enter user name: ")
    phone_number = input("Enter user phone number: ")
    users[name] = phone_number
    home()

def update_user_info():
    name = input("Enter the user name to update: ")
    if name in users.keys():
        phone_number = input("Enter user phone number: ")
        users[name] = phone_number
        home()
    else:
        print("This user does not exist in the user list")
        update_user_info()

def delete_user():
    name = input("Enter the user name to delete: ")
    if name in users.keys():
        del users[name]
        home()
    else:
        print("This user does not exist in the user list")
        delete_user()

def search_user():
    name = input("Enter the user name to search: ")
    if name in users.keys():
        print(f"User name: {name}\nUser phone number: {users[name]}")
        home()
    else:
        print("This user does not exist in the user list")
        search_user()

def output():
    print("The process is complete 😊")


def main(command):
    command = int(command)
    if command == 1:
        print(users)
        home()
    elif command == 2:
        add_user()
    elif command == 3:
        update_user_info()
    elif command == 4:
        delete_user()
    elif command == 5:
        search_user()
    elif command == 0:
        output()
    else:
        print("Not found command!!!")
        home()

home()

###################
# Desktop Notifier
import time
from plyer import notification

if __name__ == '__main__':
    while True:
        notification.notify(
            title = "ALERT!",
            message = "Take a break! It has been 20 seconds!",
            timeout = 5
        )
        time.sleep(20)
        
        
##########
from collections import defaultdict, Counter

words = ['banana', 'apple', 'cherry', 'apple']
word_counts = defaultdict(int)
for word in words:
    word_counts[word] += 1

print(word_counts)

letter_counts = Counter('banana')
print(letter_counts)

def def_value():
    return "Not Present"

d = defaultdict(def_value)
d["a"] = 1
d["b"] = 2

print(d["a"])
print(d["b"])
print(d["c"])

######################
# deque
from collections import deque

lst = [1, 2, 3, 4, 5]
de = deque(lst)

de.append(6)
de.append(7)
de.appendleft(0)
de.appendleft(-1)

de.pop()
de.popleft()

print("deque:", de)

#################
# is palindrome
def is_palindrome(string):
    new_string = ""
    for i in range(len(string)):
        if string[i] != " ":
            new_string += string[i]
    string = new_string
    for i in range(len(string)):
        if string[i].lower() != string[len(string) - i - 1].lower():
            return False
    return True



def is_palindrome(string):
    string = string.replace(" ", "").lower()
    return string == string[::-1]

#################
# Search query in Django

  # views.py faylga quyidagi funksiyani yozamiz
  from django.db.models import Q
  from django.shortcuts import render
  from myapp.models import MyModel

  def my_view(request):
    query = request.GET.get('q')
    
    if query:
    	instances = MyModel.objects.filter(
        	Q(title__icontains=query) | Q(description__icontains=query)
        )
    else:
      	instances = MyModel.objects.none()
        
    context = {'instances': instances}
    return render(request, 'search_results.html', context)

  
  # index.html da GET usuli bilan forma:
  <form method="get">
    <input type="search" name="q" />
    <button type="submit">Search</button>
  </form>

  
  # So'rovning search.html sahifada chiqishi:
  {% if instances %}
    <ul>
      {% for instance in instances %}
          <li>{{ instance.title }}</li>
      {% endfor %}
    </ul>
  {% else %}
      <p>No instances found.</p>
  {% endif %}
  
 #####################
def f(x):
    def f1(a, b):
        print("Hello")
        if b == 0:
            print("No")
            return
        return f(a, b)
    return f1

@f
def f(a, b):
    return a % b

f(6, 0)

####################
from dataclasses import dataclass

@dataclass
class User:
    name: str
    age: int
    job: str

    def user_info(self, other):
        return f"Hodim ma'lumotlari:\nIsmi: {self.name}\nYoshi: {self.age}\nKasbi: {self.job}\nOyligi: {other}"

erali = User("Erali", 23, "Dasturchi")
print(erali.name)
print(erali.user_info("$500"))

##############################
nums = [4, -2, 5, 8, -3, 0]
print("Number list: ", nums)


# The primitive way
def count_negatives(nums):
    n_negatives = 0
    for num in nums:
        if num < 0:
            n_negatives += 1
    return n_negatives
print(f"Number of negative values: {count_negatives(nums)}")


# List Comprehension - Traditional way
def count_negatives(nums):
    return len([num for num in nums if num < 0])
print(f"Number of negative values: {count_negatives(nums)}")


# List Comprehension - Tricky way
def count_negatives(nums):
    return sum([num < 0 for num in nums])
print(f"Number of negative values: {count_negatives(nums)}")

#############################
import pendulum

dt = pendulum.datetime(2023, 5, 11, tz='Asia/Tashkent')
print(dt)  # 2023-05-11T00:00:00+05:00

dt = dt.add(days=1).subtract(hours=2)

formatted = dt.format('YYYY-MM-DD HH:mm:ss')
print(formatted)  # 2023-05-11 22:00:00

###########################
# try block

from django.http import Http404
from .models import Student

def student_detail(request, pk):
    try:
        student = Student.objects.get(id=pk)
    except Student.DoesNotExist:
        raise Http404("Student doest not exists")

        # ...

####################################################

# get_object_or_404

from django.shortcuts import get_object_or_404
from .models import Student


def student_detail(request, pk):
    student = get_object_or_404(Student, id=pk)

    # ...

# Product model
class Product(models.Model):
    # relations
    category = models.ForeignKey(Category, on_delete=models.CASCADE)
    # fields
    name = models.CharField(max_length=200)
    description = models.TextField()
    slug = models.SlugField(null=True)
    image = models.ImageField(upload_to='product/%Y/%m/%d')
    properties = RichTextField()
    catalog = models.FileField(upload_to='product_catalog/', null=True)
    created = models.DateTimeField(auto_now_add=True)

    def save(self, *args, **kwargs):
        if not self.slug:
            self.slug = slugify(self.name)
        super().save(*args, **kwargs)
#################
# pip install emoji
  import emoji

  # find an emoji by emoji name
  print(emoji.emojize('Python is :thumbs_up:'))
  # Python is 👍
  print(emoji.emojize('This :telephone:very nice :smiling_face_with_smiling_eyes:'))
  # This ☎️ very nice 😊

  # find an emoji name by emoji
  print(emoji.demojize('Python is 👍'))
  # Python is :thumbs_up:

  # check emoji
  print(emoji.is_emoji('👍'))
  # True
##############
